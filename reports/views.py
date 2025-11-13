from datetime import datetime
import os, json, logging
from io import BytesIO

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Max
from django.http import JsonResponse, FileResponse
from django.shortcuts import get_object_or_404, render
from django.utils import dateformat, timezone
from django.views.decorators.http import require_http_methods
from django.db.models.functions import Lower

from docxtpl import DocxTemplate, InlineImage
from docx.shared import Mm
from openpyxl import load_workbook

from .models import Report, ReportPhoto


def _parse_date_or_today(s: str):
    """
    Принимает строку вида 'YYYY-MM-DD' и возвращает date.
    Если строка пустая/невалидная — вернёт локальную сегодняшнюю дату
    с учётом TIME_ZONE и USE_TZ.
    """
    if s:
        try:
            # строго ожидаем '2025-11-11'
            return datetime.strptime(s, "%Y-%m-%d").date()
        except ValueError:
            pass
    # локальная дата в таймзоне Django (Asia/Bishkek)
    return timezone.localdate()


def index(request):
    return render(request, 'index.html', {})


@login_required
def reports_list(request):
    if request.method == 'POST':
        try:
            pk = request.POST.get('report_id') or None

            report_data = json.loads(request.POST.get("report_data_text") or "{}")
            vehicle_data = json.loads(request.POST.get("vehicle_data_text") or "{}")
            inspection_text = json.loads(request.POST.get("inspection_text") or "{}")

            services = request.POST.get("services_table")
            materials = request.POST.get("materials_table")
            parts = request.POST.get("parts_table")
            uts = request.POST.get("uts_table")
            ost = request.POST.get("ost_table")
            phone_contacts = request.POST.get("phone_contacts", "")
            ui_state_raw = request.POST.get("ui_state")
            ui_state = json.loads(ui_state_raw) if ui_state_raw else {}

            defaults = dict(
                doc_type=report_data.get("doctype"),
                inspection_date=report_data.get("idate") or None,
                calculation_date=report_data.get("cdate") or None,
                report_number=report_data.get("contractnum"),
                ass_reason=report_data.get("assreason"),
                client_name=report_data.get("customer"),
                ownership_identification=report_data.get("property"),
                inspection_place=report_data.get("position"),
                evaluation_purpose=report_data.get("purpose"),
                results_purpose=report_data.get("appointment"),
                cost_type=report_data.get("costtype"),
                contract_price=report_data.get("contractcost"),
                contract_price_in_words=report_data.get("costinwords"),
                exchange_rate=report_data.get("exchangerate"),

                ass_object=vehicle_data.get("assobject"),
                vehicle_model=vehicle_data.get("model"),
                vehicle_year=vehicle_data.get("year"),
                vehicle_regnum=vehicle_data.get("regnum"),
                vehicle_vin=vehicle_data.get("vin"),
                vehicle_frame=vehicle_data.get("frame"),
                vehicle_passport=vehicle_data.get("passport"),
                vehicle_volume=vehicle_data.get("volume"),
                vehicle_mileage=vehicle_data.get("mileage"),
                vehicle_color=vehicle_data.get("color"),
                vehicle_type=vehicle_data.get("type"),
                vehicle_body=vehicle_data.get("body"),
                vehicle_gearbox=vehicle_data.get("gearbox"),
                vehicle_steering=vehicle_data.get("steering"),
                hourcost=vehicle_data.get("hourcost"),
                vehicle_owner=vehicle_data.get("owner"),
                vehicle_adress=vehicle_data.get("adress"),

                services_table=services,
                materials_table=materials,
                parts_table=parts,
                uts_table=uts,
                ost_table=ost,

                inspection_text_json=inspection_text,
                report_data_json=report_data,
                vehicle_data_json=vehicle_data,

                definition_text=inspection_text.get("definition", ""),
                disassembly_text=inspection_text.get("disassembly", ""),
                repair_text=inspection_text.get("repair", ""),
                painting_text=inspection_text.get("painting", ""),
                additional_text=inspection_text.get("additional", ""),
                hidden_text=inspection_text.get("hidden", ""),
                parts_text=inspection_text.get("parts", ""),
                damaged_body_parts_text=inspection_text.get("damagedBodyParts", ""),
                damaged_other_parts_text=inspection_text.get("damagedOtherParts", ""),
                unbroken_parts_text=inspection_text.get("unbrokenParts", ""),

                services_result=report_data.get("servicesres", ""),
                materials_result=report_data.get("materialsres", ""),
                total_result=report_data.get("totalres", ""),
                uts_percent=report_data.get("utspercent", ""),
                ost_percent=report_data.get("ostpercent", ""),

                kv=vehicle_data.get("kv", ""),
                kz=vehicle_data.get("kz", ""),
                kop=vehicle_data.get("kop", ""),

                phone_contacts=phone_contacts,

                ui_state=ui_state,
            )

            for key in ('inspection_date', 'calculation_date'):
                if defaults[key]:
                    try:
                        defaults[key] = datetime.strptime(defaults[key], "%Y-%m-%d").date()
                    except Exception:
                        defaults[key] = None

            if pk:
                obj = get_object_or_404(Report, pk=pk)
                for k, v in defaults.items():
                    setattr(obj, k, v)
                obj.save()
            else:
                obj = Report.objects.create(**defaults)

            return JsonResponse({'ok': True, 'pk': obj.pk})

        except Exception as e:
            return JsonResponse({'ok': False, 'error': str(e)}, status=400)

    # GET — список как был
    reports = Report.objects.order_by('inspection_date')
    num_reports = Report.objects.count()
    return render(request, 'list.html', {'num_reports': num_reports, 'reports': reports})


def is_json(myjson):
    try:
        myvar = json.loads(myjson)
    except ValueError as e:
        myvar = json.loads('[]')
    finally:
        return myvar


@login_required
def report_create(request, pk):
    all_report = get_object_or_404(Report, pk=pk)

    services_table_list = is_json(all_report.services_table)
    materials_table_list = is_json(all_report.materials_table)
    parts_table_list = is_json(all_report.parts_table)

    # --- Нормализация "запчастей" для DOCX: без копеек, т.к. шаблон подставляет ,00 ---
    def _to_float(x):
        if x is None:
            return 0.0
        if isinstance(x, (int, float)):
            return float(x)
        return float(str(x).replace(' ', '').replace(',', '.') or 0)

    def _fmt_int(n):
        return f"{int(round(n)):,.0f}".replace(",", " ")

    def _fmt_qty(q):
        """Количество без дробной части: 2.0 -> '2', 3 -> '3'.
        Если когда-то попадётся нецелое, аккуратно уберём хвостовые нули."""
        qf = float(q or 0)
        if abs(qf - round(qf)) < 1e-9:
            return str(int(round(qf)))
        # на всякий случай (если количествo бывает 0.5 и т.п.)
        return str(qf).rstrip('0').rstrip('.').replace('.', ',')

    parts_rows_docx = []
    parts_total = 0.0

    for it in (parts_table_list or []):
        name = str(it.get('text') or '')
        qty = _to_float(it.get('quant', 0))
        price = _to_float(it.get('price', it.get('norm', 0)))
        # если есть готовая сумма — берём её, иначе qty*price
        summed = it.get('sum', None)
        cost = _to_float(summed) if summed is not None else (qty * price)

        parts_total += cost

        parts_rows_docx.append({
            'text': name,
            'quant': _fmt_qty(qty),
            'norm': _fmt_int(price),  # <— БЕЗ копеек
            'cost': _fmt_int(cost),  # <— БЕЗ копеек
        })

    p_result = _fmt_int(parts_total)  # итог тоже БЕЗ копеек

    # --- Итог ремонта = Услуги + Материалы + Запчасти (без копеек, пробелы между тысячами) ---
    def _parse_money(s):
        # Приходит как "34 458" или "34 458,00" — приводим к float
        if s is None:
            return 0.0
        return float(str(s).replace(' ', '').replace(',', '.') or 0)

    s_num = _parse_money(all_report.services_result)
    m_num = _parse_money(all_report.materials_result)
    p_num = parts_total  # это уже float из блока запчастей

    svm_total_num = round(s_num + m_num + p_num)

    svm_total = _fmt_int(svm_total_num)

    # --- Сумма словами (рус.) без "сом", заглавной буквой отдельно добавим "сом" в шаблоне
    ONES_M = ['ноль', 'один', 'два', 'три', 'четыре', 'пять', 'шесть', 'семь', 'восемь', 'девять']
    ONES_F = ['ноль', 'одна', 'две', 'три', 'четыре', 'пять', 'шесть', 'семь', 'восемь', 'девять']
    TEENS = ['десять', 'одиннадцать', 'двенадцать', 'тринадцать', 'четырнадцать', 'пятнадцать',
             'шестнадцать', 'семнадцать', 'восемнадцать', 'девятнадцать']
    TENS = ['', 'десять', 'двадцать', 'тридцать', 'сорок', 'пятьдесят', 'шестьдесят', 'семьдесят', 'восемьдесят',
            'девяносто']
    HUND = ['', 'сто', 'двести', 'триста', 'четыреста', 'пятьсот', 'шестьсот', 'семьсот', 'восемьсот', 'девятьсот']

    def _morph(n, forms):  # ['сом','сома','сомов'] / ['тысяча','тысячи','тысяч']
        n = abs(n) % 100
        if 10 < n < 20: return forms[2]
        n = n % 10
        if n == 1: return forms[0]
        if 2 <= n <= 4: return forms[1]
        return forms[2]

    def _triad_to_words(n, female=False):
        """Преобразует триаду (0–999) в слова."""
        a = []
        a.append(HUND[n // 100])
        t = n % 100
        if 10 <= t <= 19:
            a.append(TEENS[t - 10])
        else:
            a.append(TENS[t // 10])
            u = n % 10
            if u > 0:  # <-- исправление: добавляем единицы только если > 0
                a.append((ONES_F if female else ONES_M)[u])
        return " ".join(filter(None, a)).strip()

    def num_to_words_ru(n):
        if n == 0: return 'ноль'
        parts = []
        # миллионы
        m = (n // 1_000_000) % 1000
        if m:
            parts += [_triad_to_words(m, False), _morph(m, ['миллион', 'миллиона', 'миллионов'])]
        # тысячи
        t = (n // 1_000) % 1000
        if t:
            parts += [_triad_to_words(t, True), _morph(t, ['тысяча', 'тысячи', 'тысяч'])]
        # единицы
        r = n % 1000
        if r:
            parts += [_triad_to_words(r, False)]
        return " ".join(parts).replace('  ', ' ').strip()

    svm_total_words = num_to_words_ru(int(svm_total_num))
    svm_total_words_cap = svm_total_words[:1].upper() + svm_total_words[1:]

    uts_table_list = is_json(all_report.uts_table)
    ost_table_list = is_json(all_report.ost_table)

    obj_lowercase = all_report.ass_object.lower()
    doc = DocxTemplate(settings.MEDIA_ROOT + "/report.docx")
    context = {'document_type': all_report.doc_type, 'assessment_reason': all_report.ass_reason,
               'used_methods': Report.METHOD_CHOICES[int(all_report.used_methods)][1],
               'assessment_object': all_report.ass_object, 'vehicle_gearbox': all_report.vehicle_gearbox,
               'vehicle_steering': all_report.vehicle_steering, 'exchange_rate': all_report.exchange_rate,
               'object_lowercase': obj_lowercase, 'contract_number': all_report.report_number,
               'inspection_date': dateformat.format(all_report.inspection_date, settings.DATE_FORMAT) + ' г.',
               'calc_date': dateformat.format(all_report.calculation_date, settings.DATE_FORMAT) + ' г.',
               'customer_name': all_report.client_name,
               'property_owner': Report.OWNERSHIP_CHOICES[int(all_report.ownership_identification)][1],
               'vehicle_location': all_report.inspection_place,
               'evaluation_purpose': Report.EVALUATION_CHOICES[int(all_report.evaluation_purpose)][1],
               'evaluation_appointment': Report.RESULTS_CHOICES[int(all_report.results_purpose)][1],
               'cost_type': Report.COST_CHOICES[int(all_report.cost_type)][1], 'cost_type_id': all_report.cost_type,
               'contract_cost': all_report.contract_price, 'vehicle_model': all_report.vehicle_model,
               'vehicle_year': all_report.vehicle_year, 'vehicle_number': all_report.vehicle_regnum,
               'vehicle_vin': all_report.vehicle_vin, 'vehicle_frame': all_report.vehicle_frame,
               'vehicle_passport': all_report.vehicle_passport, 'vehicle_engine_volume': all_report.vehicle_volume,
               'vehicle_mileage': all_report.vehicle_mileage, 'vehicle_color': all_report.vehicle_color,
               'vehicle_type': all_report.vehicle_type, 'vehicle_body_type': all_report.vehicle_body,
               'cost_per_hour': all_report.hourcost, 'vehicle_owner': all_report.vehicle_owner,
               'vehicle_adress': all_report.vehicle_adress, 'disassembly': all_report.disassembly_text,
               'damagedBodyParts': all_report.damaged_body_parts_text, 'unbrokenParts': all_report.unbroken_parts_text,
               'damagedOtherParts': all_report.damaged_other_parts_text, 'definition': all_report.definition_text,
               'repair': all_report.repair_text, 'painting': all_report.painting_text,
               'additional': all_report.additional_text, 'hidden': all_report.hidden_text,
               'parts': all_report.parts_text, 'services_table': services_table_list,
               'phone_contacts': all_report.phone_contacts,
               'materials_table': materials_table_list, 'parts_table': parts_table_list, 'uts_table': uts_table_list,
               'ost_table': ost_table_list, 's_result': all_report.services_result,
               'm_result': all_report.materials_result, 't_result': all_report.total_result,
               'uts_percent': all_report.uts_percent, 'ost_percent': all_report.ost_percent, 'kv': all_report.kv,
               'kz': all_report.kz, 'kop': all_report.kop, 'parts_table': parts_rows_docx, 'p_result': p_result,
               'svm_total': svm_total, 'svm_total_words_cap': svm_total_words_cap, }

    if all_report.cost_type == "0":
        file_name = ' Ремонт.docx'
    elif all_report.cost_type == "1":
        file_name = ' Ремонт и УТС.docx'
    elif all_report.cost_type == "2":
        file_name = ' Лом.docx'
    else:
        file_name = ' Лом и ремонт.docx'

    photos_qs = (ReportPhoto.objects
                 .filter(report_id=pk)
                 .order_by('order', 'id'))

    photos_qs = (ReportPhoto.objects
                 .filter(report_id=pk)
                 .order_by('order', 'id'))

    photo_pairs = []
    pair = []
    counter = 1  # счётчик для нумерации фотографий

    for p in photos_qs:
        item = {
            'num': counter,
            'img': InlineImage(doc, p.image.path, width=Mm(80)),  # ширину подгони под свой шаблон
            'cap': p.caption or ''
        }
        counter += 1
        pair.append(item)
        if len(pair) == 2:
            photo_pairs.append({'left': pair[0], 'right': pair[1]})
            pair = []

    # если нечётное — завершим последним «пустым правым»
    if len(pair) == 1:
        photo_pairs.append({
            'left': pair[0],
            'right': {'num': '', 'img': '', 'cap': ''}
        })

    context.update({
        'has_photos': photos_qs.exists(),
        'photo_pairs': photo_pairs,
    })

    doc.render(context)
    byte_io = BytesIO()
    doc.save(byte_io)
    byte_io.seek(0)
    return FileResponse(byte_io, as_attachment=True, filename=all_report.report_number.split('/')[0] + file_name)


def contract_create(request, pk):
    all_report = get_object_or_404(Report, pk=pk)
    doc = DocxTemplate(settings.MEDIA_ROOT + "/contract.docx")
    context = {'document_type': all_report.doc_type, 'assessment_object': all_report.ass_object,
               'contract_number': all_report.report_number,
               'inspection_date': dateformat.format(all_report.inspection_date, settings.DATE_FORMAT) + ' г.',
               'customer_name': all_report.client_name,
               'evaluation_purpose': Report.EVALUATION_CHOICES[int(all_report.evaluation_purpose)][1],
               'cost_type': Report.COST_CHOICES[int(all_report.cost_type)][1],
               'contract_cost': all_report.contract_price, 'contract_cost_in_words': all_report.contract_price_in_words,
               'vehicle_model': all_report.vehicle_model, 'vehicle_number': all_report.vehicle_regnum, }
    doc.render(context)
    byte_io = BytesIO()
    doc.save(byte_io)
    byte_io.seek(0)
    return FileResponse(byte_io, as_attachment=True, filename=all_report.report_number.split('/')[0] + " Договор.docx")


def cash_document(request, pk):
    all_report = get_object_or_404(Report, pk=pk)
    services_table_list = is_json(all_report.services_table)
    materials_table_list = is_json(all_report.materials_table)
    parts_table_list = is_json(all_report.parts_table)
    uts_table_list = is_json(all_report.uts_table)
    ost_table_list = is_json(all_report.ost_table)
    wb = load_workbook(settings.MEDIA_ROOT + "/excel.xlsx")
    activeSheet = wb.active
    activeSheet["B1"] = all_report.doc_type
    activeSheet["B2"] = all_report.report_number
    activeSheet["B3"] = all_report.ass_reason
    activeSheet["B4"] = dateformat.format(all_report.inspection_date, settings.DATE_FORMAT) + ' г.'
    activeSheet["B5"] = dateformat.format(all_report.calculation_date, settings.DATE_FORMAT) + ' г.'
    activeSheet["B6"] = all_report.client_name
    activeSheet["B7"] = Report.OWNERSHIP_CHOICES[int(all_report.ownership_identification)][1]
    activeSheet["B8"] = all_report.inspection_place
    activeSheet["B9"] = Report.EVALUATION_CHOICES[int(all_report.evaluation_purpose)][1]
    activeSheet["B10"] = Report.RESULTS_CHOICES[int(all_report.results_purpose)][1]
    activeSheet["B11"] = Report.COST_CHOICES[int(all_report.cost_type)][1]
    activeSheet["B12"] = Report.METHOD_CHOICES[int(all_report.used_methods)][1]
    activeSheet["B13"] = float(all_report.contract_price.replace(' ', '').replace(',', '.'))
    activeSheet["B14"] = all_report.contract_price_in_words

    activeSheet["B16"] = all_report.ass_object
    activeSheet["B17"] = all_report.vehicle_model
    activeSheet["B18"] = all_report.vehicle_regnum
    activeSheet["B19"] = all_report.vehicle_vin
    activeSheet["B20"] = all_report.vehicle_frame
    activeSheet["B21"] = all_report.vehicle_passport
    activeSheet["B22"] = all_report.vehicle_year
    activeSheet["B23"] = all_report.vehicle_volume
    activeSheet["B24"] = all_report.vehicle_mileage
    activeSheet["B25"] = all_report.vehicle_color
    activeSheet["B26"] = all_report.vehicle_type + ' ' + all_report.vehicle_body
    activeSheet["B27"] = all_report.vehicle_gearbox
    activeSheet["B28"] = all_report.vehicle_steering
    activeSheet["B29"] = all_report.vehicle_owner
    activeSheet["B30"] = all_report.vehicle_adress

    activeSheet["B32"] = float(all_report.exchange_rate)
    activeSheet["B33"] = all_report.hourcost

    utsSheet = wb.worksheets[7]
    utsLen = len(uts_table_list)
    for elem in range(utsLen):
        utsSheet.cell(column=1, row=elem + 2, value=elem + 1)
        utsSheet.cell(column=2, row=elem + 2, value=uts_table_list[elem]['text'])
        utsSheet.cell(column=3, row=elem + 2, value=uts_table_list[elem]['quant'])
        utsSheet.cell(column=4, row=elem + 2,
                      value=float(uts_table_list[elem]['uts'].replace(' ', '').replace(',', '.')))
        utsSheet.cell(column=5, row=elem + 2, value='=D{}*Средн!$D$22/100'.format(elem + 2))
    utsSheet["G2"] = '=SUM(D2:D{}'.format(utsLen + 1) + ')'
    utsSheet["H2"] = '=SUM(E2:E{}'.format(utsLen + 1) + ')'
    utsSheet["G4"] = float(all_report.uts_percent.replace(' ', '').replace(',', '.'))

    ostSheet = wb.worksheets[8]
    ostLen = len(ost_table_list)
    for elem in range(ostLen):
        ostSheet.cell(column=1, row=elem + 3, value=ost_table_list[elem]['text'])
        ostSheet.cell(column=2, row=elem + 3,
                      value=float(ost_table_list[elem]['ost'].replace(' ', '').replace(',', '.')))
        ostSheet.cell(column=3, row=elem + 3, value='=($E$2*$F$2*$G$2*Средн!$D$22*B{})/100'.format(elem + 3))
    ostSheet["E2"] = float(all_report.kz.replace(' ', '').replace(',', '.'))
    ostSheet["F2"] = float(all_report.kv.replace(' ', '').replace(',', '.'))
    ostSheet["G2"] = float(all_report.kop.replace(' ', '').replace(',', '.'))
    ostSheet["E4"] = float(all_report.ost_percent.replace(' ', '').replace(',', '.'))
    ostSheet["F4"] = '=SUM(B3:B{}'.format(ostLen + 2) + ')'
    ostSheet["G4"] = '=SUM(C3:C{}'.format(ostLen + 2) + ')'

    damageSheet = wb.worksheets[5]
    servicesLen = len(services_table_list)
    for elem in range(servicesLen):
        damageSheet.cell(column=1, row=elem + 2, value=elem + 1)
        damageSheet.cell(column=2, row=elem + 2, value=services_table_list[elem]['text'])
        damageSheet.cell(column=3, row=elem + 2, value=services_table_list[elem]['quant'])
        damageSheet.cell(column=4, row=elem + 2,
                         value=float(services_table_list[elem]['norm'].replace(' ', '').replace(',', '.')))
        damageSheet.cell(column=5, row=elem + 2, value='=C{}'.format(elem + 2) + '*D{}*Дано!$B$33'.format(elem + 2))
    damageSheet["G2"] = '=SUM(E2:E{}'.format(servicesLen + 1) + ')'
    materialsLen = len(materials_table_list)
    materialsStartRow = servicesLen + 4
    damageSheet.cell(column=1, row=materialsStartRow - 1, value='№')
    damageSheet.cell(column=2, row=materialsStartRow - 1, value='Материалы')
    damageSheet.cell(column=3, row=materialsStartRow - 1, value='Кол-во')
    damageSheet.cell(column=4, row=materialsStartRow - 1, value='Цена')
    damageSheet.cell(column=5, row=materialsStartRow - 1, value='Стоимость')
    for elem in range(materialsLen):
        damageSheet.cell(column=1, row=elem + materialsStartRow, value=elem + 1)
        damageSheet.cell(column=2, row=elem + materialsStartRow, value=materials_table_list[elem]['text'])
        damageSheet.cell(column=3, row=elem + materialsStartRow, value=materials_table_list[elem]['quant'])
        damageSheet.cell(column=4, row=elem + materialsStartRow,
                         value=float(materials_table_list[elem]['norm'].replace(' ', '').replace(',', '.')))
        damageSheet.cell(column=5, row=elem + materialsStartRow,
                         value='=C{}'.format(elem + materialsStartRow) + '*D{}'.format(elem + materialsStartRow))
    partsLen = len(parts_table_list)
    partsStartRow = materialsLen + materialsStartRow + 2
    damageSheet["H2"] = '=SUM(E{}:E{})'.format(materialsStartRow, partsStartRow - 3)

    # Заголовки (как были)
    damageSheet.cell(column=1, row=partsStartRow - 1, value='№')
    damageSheet.cell(column=2, row=partsStartRow - 1, value='Запасные части')
    damageSheet.cell(column=3, row=partsStartRow - 1, value='Кол-во')
    damageSheet.cell(column=4, row=partsStartRow - 1, value='Цена')
    damageSheet.cell(column=5, row=partsStartRow - 1, value='Стоимость')

    for elem in range(partsLen):
        row = elem + partsStartRow
        name = parts_table_list[elem].get('text', '')
        qty = parts_table_list[elem].get('quant', 0)
        price = parts_table_list[elem].get('price', parts_table_list[elem].get('norm', 0))

        # Числа приводим к float
        def _f(v):
            try:
                return float(str(v).replace(' ', '').replace(',', '.') or 0)
            except Exception:
                return 0.0

        damageSheet.cell(column=1, row=row, value=elem + 1)
        damageSheet.cell(column=2, row=row, value=name)
        damageSheet.cell(column=3, row=row, value=_f(qty))
        damageSheet.cell(column=4, row=row, value=_f(price))  # <-- РЕАЛЬНАЯ ЦЕНА
        damageSheet.cell(column=5, row=row, value='=C{}*D{}'.format(row, row))  # Стоимость = C*D

    if partsLen > 0:
        damageSheet["I2"] = '=SUM(E{}:E{})'.format(partsStartRow, partsStartRow + partsLen - 1)
    else:
        damageSheet["I2"] = 0
    damageSheet["G4"] = float(all_report.services_result.replace(' ', '').replace(',', '.'))
    damageSheet["H4"] = float(all_report.materials_result.replace(' ', '').replace(',', '.'))
    byte_io = BytesIO()
    wb.save(byte_io)
    byte_io.seek(0)
    return FileResponse(byte_io, as_attachment=True, filename=all_report.report_number.split('/')[0] + ".xlsx")


def report_json(request, pk: int):
    """Отдаёт JSON для префилла UI по отчёту pk."""
    report = get_object_or_404(Report, pk=pk)

    def d(val):
        return "" if val is None else str(val)

    data = {
        "fields": {
            # раздел «Данные отчёта»
            "doc_type": d(report.doc_type),
            "inspection_date": report.inspection_date.strftime("%Y-%m-%d") if report.inspection_date else "",
            "calc_date": report.calculation_date.strftime("%Y-%m-%d") if report.calculation_date else "",
            "contract_number": d(report.report_number),
            "ass_reason": d(report.ass_reason),
            "customer_name": d(report.client_name),
            "property_owner": d(report.ownership_identification),  # value у <select>
            "vehicle_location": d(report.inspection_place),
            "evaluation_purpose": d(report.evaluation_purpose),
            "evaluation_appointment": d(report.results_purpose),
            "cost_type": d(report.cost_type),
            "contract_cost": d(report.contract_price),
            "contract_cost_in_words": d(report.contract_price_in_words),
            "exchange_rate": d(report.exchange_rate),

            # раздел «Данные объекта оценки»
            "ass_object": d(report.ass_object),
            "vehicle_model": d(report.vehicle_model),
            "vehicle_year": d(report.vehicle_year),
            "vehicle_number": d(report.vehicle_regnum),
            "vehicle_vin": d(report.vehicle_vin),
            "vehicle_frame": d(report.vehicle_frame),
            "vehicle_passport": d(report.vehicle_passport),
            "vehicle_engine_volume": d(report.vehicle_volume),
            "vehicle_mileage": d(report.vehicle_mileage),
            "vehicle_color": d(report.vehicle_color),
            "vehicle_type": d(report.vehicle_type),  # текстовое поле рядом с селектом типов
            "vehicle_body_type": d(report.vehicle_body),  # текстовое поле рядом с селектом кузова
            "vehicle_gearbox": d(report.vehicle_gearbox),
            "vehicle_steering": d(report.vehicle_steering),
            "cost_per_hour": d(report.hourcost),
            "class": d(report.hourcost),  # чтобы синхронизировать селект «Класс ТС»
            "vehicle_owner": d(report.vehicle_owner),
            "vehicle_adress": d(report.vehicle_adress),
            "phone_contacts": d(report.phone_contacts),
        },

        # Блоки работ/материалов (чекбоксы и т.д.) восстанавливаем, если вы начнёте хранить «снимок UI».
        # На первом этапе оставьте пустым — см. Шаг 5 (опционально) ниже.
        "blocks": report.ui_state.get("blocks", []) if report.ui_state else [],
        "parts_table": json.loads(report.parts_table) if report.parts_table else [],

    }
    return JsonResponse(data)


def edit(request, pk: int):
    # Рендерим ту же index.html, но прокидываем pk
    return render(request, 'index.html', {"prefill_id": pk})


@login_required
def list_photos(request, pk):
    photos = (ReportPhoto.objects
              .filter(report_id=pk)
              .order_by(Lower('caption'), 'order', 'id'))
    data = [{
        'id': p.id,
        'url': p.image.url,
        'caption': p.caption or '',
        'order': p.order or 0,
    } for p in photos]
    return JsonResponse({'photos': data})


@login_required
@require_http_methods(["GET"])
def photos_json(request, report_id: int):
    """Список фото отчёта для префилла (edit)."""
    report = get_object_or_404(Report, pk=report_id)
    qs = ReportPhoto.objects.filter(report=report).order_by('order', 'id')
    photos = [{
        'id': p.id,
        'url': request.build_absolute_uri(p.image.url),
        'caption': p.caption or '',
        'order': p.order or 0,
    } for p in qs]
    return JsonResponse({'ok': True, 'photos': photos})


@login_required
@require_http_methods(["POST"])
@transaction.atomic
def photos_upload(request, report_id: int):
    """
    Синхронизация фото отчёта.

    РЕЖИМ A (современный, безопасный): mode=delta
      • Удаляем ТОЛЬКО id из removed_ids.
      • Обновляем подписи оставшимся (captions_by_id).
      • Обновляем порядок order по order_map.
      • Дозагружаем новые (images[] + captions_new).
      • Ничего не «чистим» по умолчанию.

    РЕЖИМ B (совместимость со старым фронтом): без mode=delta
      • Если keep_ids пуст/бит ('' | [] | null | None | undefined) → FAIL-SAFE: НЕ удаляем никого.
      • Иначе удаляем тех, кого нет в keep_ids, обновляем подписи у оставшихся, обновляем порядок.
    """
    log = logging.getLogger(__name__)
    report = get_object_or_404(Report, pk=report_id)

    # --- Текущие фото и их id ---
    existing_qs = ReportPhoto.objects.filter(report=report).order_by('order', 'id')
    existing_ids = list(existing_qs.values_list('id', flat=True))

    mode = (request.POST.get('mode') or '').strip().lower()

    # --- Разбор json-полей ---
    def parse_json_dict(s, default):
        try:
            obj = json.loads(s) if s is not None else default
            return obj if isinstance(obj, dict) else default
        except Exception:
            return default

    def parse_json_list(s, default):
        try:
            obj = json.loads(s) if s is not None else default
            return obj if isinstance(obj, list) else default
        except Exception:
            return default

    captions_by_id = parse_json_dict(request.POST.get('captions_by_id'), {})
    captions_new = parse_json_list(request.POST.get('captions_new'), [])
    order_map = parse_json_dict(request.POST.get('order_map'), {})  # <-- карта порядка

    files = request.FILES.getlist('images[]')

    if mode == 'delta':
        # ===== РЕЖИМ DELTA =====
        removed_raw = (request.POST.get('removed_ids') or '').strip()
        if removed_raw in ('', '[]', 'null', 'None', 'undefined'):
            removed_ids = []
        else:
            removed_ids = [int(x) for x in removed_raw.split(',') if x.strip().isdigit()]

        # Удаляем только removed_ids
        if removed_ids:
            doomed = ReportPhoto.objects.filter(report=report, id__in=removed_ids)
            for p in doomed:
                try:
                    if p.image and os.path.isfile(p.image.path):
                        os.remove(p.image.path)
                except Exception:
                    pass
            doomed.delete()

        # Обновляем подписи оставшимся
        if captions_by_id:
            for p in ReportPhoto.objects.filter(report=report).exclude(id__in=removed_ids):
                new_cap = captions_by_id.get(str(p.id))
                if new_cap is None:
                    new_cap = captions_by_id.get(p.id)
                if new_cap is not None:
                    new_cap = (new_cap or '')
                    if new_cap != (p.caption or ''):
                        p.caption = new_cap
                        p.save(update_fields=['caption'])

        # Обновляем порядок (order) оставшимся
        if order_map:
            for p in ReportPhoto.objects.filter(report=report).exclude(id__in=removed_ids):
                new_order = order_map.get(str(p.id))
                if new_order is None:
                    new_order = order_map.get(p.id)
                if new_order is None:
                    continue
                try:
                    new_order = int(new_order)
                except (TypeError, ValueError):
                    continue
                if new_order < 0:
                    continue
                if p.order != new_order:
                    p.order = new_order
                    p.save(update_fields=['order'])

    else:
        # ===== LEGACY: через keep_ids =====
        raw = (request.POST.get('keep_ids') or '').strip()
        EMPTY = {'', '[]', 'null', 'None', 'undefined'}

        if raw in EMPTY:
            keep_ids = existing_ids[:]  # FAIL-SAFE: никого не удаляем
        else:
            keep_ids = [int(x) for x in raw.split(',') if x.strip().isdigit()]

        # Доп. предохранитель: если keep пуст, а отчёт имел фото — не чистим
        if not keep_ids and existing_ids:
            keep_ids = existing_ids[:]

        ids_to_delete = set(existing_ids) - set(keep_ids)
        if ids_to_delete:
            doomed = ReportPhoto.objects.filter(report=report, id__in=ids_to_delete)
            for p in doomed:
                try:
                    if p.image and os.path.isfile(p.image.path):
                        os.remove(p.image.path)
                except Exception:
                    pass
            doomed.delete()

        # Подписи у оставшихся
        if keep_ids and captions_by_id:
            for p in ReportPhoto.objects.filter(report=report, id__in=keep_ids):
                new_cap = captions_by_id.get(str(p.id))
                if new_cap is None:
                    new_cap = captions_by_id.get(p.id)
                if new_cap is not None:
                    new_cap = (new_cap or '')
                    if new_cap != (p.caption or ''):
                        p.caption = new_cap
                        p.save(update_fields=['caption'])

        # Порядок у оставшихся (legacy)
        if keep_ids and order_map:
            for p in ReportPhoto.objects.filter(report=report, id__in=keep_ids):
                new_order = order_map.get(str(p.id))
                if new_order is None:
                    new_order = order_map.get(p.id)
                if new_order is None:
                    continue
                try:
                    new_order = int(new_order)
                except (TypeError, ValueError):
                    continue
                if new_order < 0:
                    continue
                if p.order != new_order:
                    p.order = new_order
                    p.save(update_fields=['order'])

    # ===== Новые фото =====
    last_order = (ReportPhoto.objects.filter(report=report)
                  .aggregate(max_order=Max('order')).get('max_order') or 0)
    seq = last_order
    for idx, f in enumerate(files):
        cap = (captions_new[idx] if idx < len(captions_new) else '') or ''
        seq += 1
        ReportPhoto.objects.create(
            report=report,
            image=f,
            caption=cap,
            order=seq
        )

    # ===== Ответ: полный актуальный список =====
    all_now = ReportPhoto.objects.filter(report=report).order_by('order', 'id')
    resp = [{
        'id': p.id,
        'url': request.build_absolute_uri(p.image.url),
        'caption': p.caption or '',
        'order': p.order or 0,
    } for p in all_now]

    return JsonResponse({
        'ok': True,
        'mode': mode or 'legacy',
        'photos': resp,
        'saved_count': len(files),
        'total': all_now.count(),
    })
